import sys
import json
from time import sleep

from bs4 import BeautifulSoup as bs
import PySimpleGUI as GUI
import openpyxl
import requests


class Parser:
    GUI.theme('LightGreen5')
    GUI.set_options(font='Franklin 10')
    WINDOW_SIZE = 1100, 650
    WINDOW_TITLE = 'Product Parser'
    DEFAULT_URL = 'https://mi-shop.com/ru/catalog/smartphones/'
    PAGE_URL_POSTFIX = 'page/'
    OUTPUT_BOX_SIZE = 120, 23
    EVENT_BOX_SIZE = 30, 23
    PARAM_INPUT_SIZE = 16, 1
    PARAM_TITLE_SIZE = 11, 1
    BUTTON_SIZE = 9, 1
    OUTPUT_TEXTBOX_PAD = 5, 10
    DROPDOWN_MENU_SIZE = 14, 1
    CONTAINER_TYPES = ['div', 'span', 'section']
    SELECTOR_TYPES = ['class', 'text', 'name', 'id']
    PARSING_DELAY_TIME = 1
    ABOUT_AUTHOR = 'Made by frostfree to help the friend of him\nGitHub: None'
    USER_AGENT_HEADER = 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36  \
                        (KHTML, like Gecko) Chrome/100.0.4896.160 YaBrowser/22.5.3.684 \
                        Yowser/2.5 Safari/537.36'

    def __init__(self):
        # Input elements layout
        menu = GUI.Menu([
            ['Configuration', ['Load::conf', 'Save::conf']],
            ['Help', ['About']]
        ])
        pages_text = GUI.Text('pages:')
        pages_input = GUI.Spin([i for i in range(1,16)], initial_value=1, size=(3,1), readonly=True, key='-PAGE_NUMBER-')
        url_input = GUI.Input(default_text=self.DEFAULT_URL, expand_x=True, key='-URL-')
        request_btn = GUI.Button('Request', size=self.BUTTON_SIZE, key='-REQUEST-')
        open_btn = GUI.Button('Open', size=self.BUTTON_SIZE, key='-OPEN-')
        show_page_btn = GUI.Button('Show Page', size=self.BUTTON_SIZE, key='-PAGE_SHOW-', disabled=True)
        
        # Searching elements layout
        find_all_btn = GUI.Button('Find all', size=self.BUTTON_SIZE, key='-FIND_ALL-', disabled=True)
        find_container = GUI.Text(text='Container type:', size=self.PARAM_TITLE_SIZE)
        find_selector_type = GUI.Text(text='Selector type:', size=self.PARAM_TITLE_SIZE)
        find_selector_name = GUI.Text(text='Selector name:', size=self.PARAM_TITLE_SIZE)
        param_container = GUI.Combo(values=self.CONTAINER_TYPES, default_value=self.CONTAINER_TYPES[0], key='-CONTAINER-', size=self.DROPDOWN_MENU_SIZE)
        param_selector_type = GUI.Combo(values=self.SELECTOR_TYPES, default_value=self.SELECTOR_TYPES[0], key='-SELECTOR_TYPE-', size=self.DROPDOWN_MENU_SIZE)
        param_selector_name = GUI.Input(default_text='product-card__body', size=self.PARAM_INPUT_SIZE, key='-SELECTOR_NAME-')
        find_column = GUI.Column([
            [find_container, param_container],
            [find_selector_type, param_selector_type],
            [find_selector_name, param_selector_name],
            [find_all_btn],
            [show_page_btn],
        ])
        # Parsing elements layout
        parse_btn = GUI.Button('Parse', key='-PARSE-', size=self.BUTTON_SIZE, disabled=True)
        parse_title = GUI.Text(text='', size=self.PARAM_TITLE_SIZE)
        parse_container = GUI.Text(text='Container type:', size=self.PARAM_TITLE_SIZE)
        parse_selector_type = GUI.Text(text='Selector type:', size=self.PARAM_TITLE_SIZE)
        parse_selector_name = GUI.Text(text='Selector name:', size=self.PARAM_TITLE_SIZE)
        parse_names = GUI.Text(text='Save as:', size=self.PARAM_TITLE_SIZE)
        param_info_column = GUI.Column([
            [parse_title],
            [parse_container],
            [parse_selector_type],
            [parse_selector_name],
            [parse_names],
            [parse_btn],
        ])
        param_1_clear = GUI.Button('Clear', key='-CLEAR_PARAM_1-', size=self.BUTTON_SIZE)
        param_1_title = GUI.Text(text='Parameter 1', size=self.PARAM_TITLE_SIZE)
        param_1_container = GUI.Combo(values=self.CONTAINER_TYPES, default_value=self.CONTAINER_TYPES[0], key='-PARAM_CONTAINER_1-', size=self.DROPDOWN_MENU_SIZE)
        param_1_selector_type = GUI.Combo(values=self.SELECTOR_TYPES, default_value=self.SELECTOR_TYPES[0], key='-PARAM_SELECTOR_TYPE_1-', size=self.DROPDOWN_MENU_SIZE)
        param_1_selector_name = GUI.Input(default_text='product-card__type', key='-PARAM_SELECTOR_NAME_1-', size=self.PARAM_INPUT_SIZE)
        param_1_name = GUI.Input(default_text='Категория',key='-PARAM_NAME_1-', size=self.PARAM_INPUT_SIZE)
        param_1_column = GUI.Column([
            [param_1_title],
            [param_1_container],
            [param_1_selector_type],
            [param_1_selector_name],
            [param_1_name],
            [param_1_clear],
        ])
        param_2_clear = GUI.Button('Clear', key='-CLEAR_PARAM_2-', size=self.BUTTON_SIZE)
        param_2_title = GUI.Text(text='Parameter 2', size=self.PARAM_TITLE_SIZE)
        param_2_container = GUI.Combo(values=self.CONTAINER_TYPES, default_value=self.CONTAINER_TYPES[0], key='-PARAM_CONTAINER_2-', size=self.DROPDOWN_MENU_SIZE)
        param_2_selector_type = GUI.Combo(values=self.SELECTOR_TYPES, default_value=self.SELECTOR_TYPES[0], key='-PARAM_SELECTOR_TYPE_2-', size=self.DROPDOWN_MENU_SIZE)
        param_2_selector_name = GUI.Input(default_text='product-card__title', key='-PARAM_SELECTOR_NAME_2-', size=self.PARAM_INPUT_SIZE)
        param_2_name = GUI.Input(default_text='Наименование',key='-PARAM_NAME_2-', size=self.PARAM_INPUT_SIZE)
        param_2_column = GUI.Column([
            [param_2_title],
            [param_2_container],
            [param_2_selector_type],
            [param_2_selector_name],
            [param_2_name],
            [param_2_clear],
        ])
        param_3_clear = GUI.Button('Clear', key='-CLEAR_PARAM_3-', size=self.BUTTON_SIZE)
        param_3_title = GUI.Text(text='Parameter 3', size=self.PARAM_TITLE_SIZE)
        param_3_container = GUI.Combo(values=self.CONTAINER_TYPES, default_value=self.CONTAINER_TYPES[1], key='-PARAM_CONTAINER_3-', size=self.DROPDOWN_MENU_SIZE)
        param_3_selector_type = GUI.Combo(values=self.SELECTOR_TYPES, default_value=self.SELECTOR_TYPES[0], key='-PARAM_SELECTOR_TYPE_3-', size=self.DROPDOWN_MENU_SIZE)
        param_3_selector_name = GUI.Input(default_text='price__new', key='-PARAM_SELECTOR_NAME_3-', size=self.PARAM_INPUT_SIZE)
        param_3_name = GUI.Input(default_text='Цена',key='-PARAM_NAME_3-', size=self.PARAM_INPUT_SIZE)
        param_3_column = GUI.Column([
            [param_3_title],
            [param_3_container],
            [param_3_selector_type],
            [param_3_selector_name],
            [param_3_name],
            [param_3_clear],
        ])
        param_4_clear = GUI.Button('Clear', key='-CLEAR_PARAM_4-', size=self.BUTTON_SIZE)
        param_4_title = GUI.Text(text='Parameter 4', size=self.PARAM_TITLE_SIZE)
        param_4_container = GUI.Combo(values=self.CONTAINER_TYPES, default_value=self.CONTAINER_TYPES[0], key='-PARAM_CONTAINER_4-', size=self.DROPDOWN_MENU_SIZE)
        param_4_selector_type = GUI.Combo(values=self.SELECTOR_TYPES, default_value=self.SELECTOR_TYPES[0], key='-PARAM_SELECTOR_TYPE_4-', size=self.DROPDOWN_MENU_SIZE)
        param_4_selector_name = GUI.Input(default_text='product-card__reviews', key='-PARAM_SELECTOR_NAME_4-', size=self.PARAM_INPUT_SIZE)
        param_4_name = GUI.Input(default_text='Популярность',key='-PARAM_NAME_4-', size=self.PARAM_INPUT_SIZE)
        param_4_column = GUI.Column([
            [param_4_title],
            [param_4_container],
            [param_4_selector_type],
            [param_4_selector_name],
            [param_4_name],
            [param_4_clear],
        ])
        param_5_clear = GUI.Button('Clear', key='-CLEAR_PARAM_5-', size=self.BUTTON_SIZE)
        param_5_title = GUI.Text(text='Parameter 5', size=self.PARAM_TITLE_SIZE)
        param_5_container = GUI.Combo(values=self.CONTAINER_TYPES, default_value=self.CONTAINER_TYPES[0], key='-PARAM_CONTAINER_5-', size=self.DROPDOWN_MENU_SIZE)
        param_5_selector_type = GUI.Combo(values=self.SELECTOR_TYPES, default_value=self.SELECTOR_TYPES[0], key='-PARAM_SELECTOR_TYPE_5-', size=self.DROPDOWN_MENU_SIZE)
        param_5_selector_name = GUI.Input(default_text='product-card__evaluation', key='-PARAM_SELECTOR_NAME_5-', size=self.PARAM_INPUT_SIZE)
        param_5_name = GUI.Input(default_text='Оценка',key='-PARAM_NAME_5-', size=self.PARAM_INPUT_SIZE)
        param_5_column = GUI.Column([
            [param_5_title],
            [param_5_container],
            [param_5_selector_type],
            [param_5_selector_name],
            [param_5_name],
            [param_5_clear],
        ])
        #  Output elements layout
        event_box = GUI.Multiline(key='-MESSAGE-', disabled=True, size=self.EVENT_BOX_SIZE, reroute_stdout=True, autoscroll=True, no_scrollbar=True)
        output_box = GUI.Multiline(key='-OUTPUT-', disabled=True, size=self.OUTPUT_BOX_SIZE)
        save_json_btn = GUI.Button('Save .json',  key='-SAVE_JSON-', size=self.BUTTON_SIZE, disabled=True)
        save_xlsx_btn = GUI.Button('Save .xlsx',  key='-SAVE_XLSX-', size=self.BUTTON_SIZE, disabled=True)
        clear_btn = GUI.Button('Clear', key='-CLEAR-', size=self.BUTTON_SIZE)
        close_btn = GUI.Button('Close', key='-CLOSE-', size=self.BUTTON_SIZE)

        layout = [
            [menu],
            [open_btn, request_btn, url_input, pages_text, pages_input],
            [GUI.HorizontalSeparator()],
            [find_column, GUI.Push(), param_info_column, param_1_column, param_2_column, param_3_column, param_4_column, param_5_column],
            [GUI.HorizontalSeparator()],
            [output_box, GUI.Push(), event_box],
            [save_json_btn, save_xlsx_btn, GUI.Push(), clear_btn, close_btn],
        ]
        # These keys shouldn't be saved in conf file
        self.not_conf_keys = [event_box.key, output_box.key]
        
        # Data storage
        self.page = None
        self.data = []
        self.clean_data = []
        self.window = GUI.Window(
            title=self.WINDOW_TITLE,
            layout=layout,
            size=self.WINDOW_SIZE,
        )

    def run(self):
        while True:
            event, values = self.window.read()

            if event in ('-OPEN-', '-REQUEST-'):
                if event == '-OPEN-':
                    path = GUI.popup_get_file('Open', no_window=True)
                    if path:
                        self.load_data(source=path)
                
                elif event == '-REQUEST-':
                    url = values['-URL-']
                    pages = values['-PAGE_NUMBER-']
                    if pages and url:
                        self.request_data(source=url, pages=pages)

                if self.page:
                    self.window['-FIND_ALL-'].update(disabled=False)
                    self.window['-PAGE_SHOW-'].update(disabled=False)


            if event == '-FIND_ALL-':
                self.find_data(
                    container=values['-CONTAINER-'],
                    attrs={values['-SELECTOR_TYPE-']: values['-SELECTOR_NAME-']},
                )
                if self.data:
                    self.window['-PARSE-'].update(disabled=False)
            
            if event == '-PAGE_SHOW-':
                if self.page:
                    self.window['-OUTPUT-'].update(self.page)
                else:
                    print('page is not loaded')

            if event == '-PARSE-':
                params = []
                stop = False
                for param_number in range(1, 6):
                    name = values[f'-PARAM_NAME_{param_number}-'].strip()
                    container = values[f'-PARAM_CONTAINER_{param_number}-'].strip()
                    selector_type = values[f'-PARAM_SELECTOR_TYPE_{param_number}-'].strip()
                    selector_name = values[f'-PARAM_SELECTOR_NAME_{param_number}-'].strip()
                    fields = [name, container, selector_type, selector_name]
                    if all(fields):
                        param = {
                            'name': name,
                            'container': container,
                            'attrs': {selector_type: selector_name},
                        }
                        params.append(param)

                    elif any(fields):
                        stop = True
                        print(f'column №{param_number} is not completed')

                if not stop:
                    self.parse_data(params)
                    if self.clean_data:
                        self.window['-SAVE_JSON-'].update(disabled=False)
                        self.window['-SAVE_XLSX-'].update(disabled=False)

            if event in [f'-CLEAR_PARAM_{i}-' for i in range(1, 6)]:
                param_number = event.split('_')[-1].replace('-','')
                self.clear_param(param_number)

            if event == 'Load::conf':
                path = GUI.popup_get_file('Load', no_window=True)
                if path:
                    self.load_conf()

            if event in ('-SAVE_JSON-', 'Save::conf'):
                if event == '-SAVE_JSON-':
                    data = self.clean_data
                elif event == 'Save::conf':
                    data = {key: value for key, value in values.items() 
                            if key not in self.not_conf_keys and isinstance(key, str)}
                
                path = GUI.popup_get_file('Save as', no_window=True, save_as=True)
                if path:
                    self.save_json(data, path)

            if event == '-SAVE_XLSX-':
                self.save_xlsx()

            if event == '-CLEAR-':
                self.clear_data()

            if event == 'About':
                GUI.popup_scrolled(self.ABOUT_AUTHOR, title='About', size=(15, 5), no_sizegrip=True)
        
            if event in (GUI.WIN_CLOSED, '-CLOSE-'):
                break

    def load_data(self, source):
        with open(source, mode='r', encoding='utf-8') as html_file:
            self.page = bs(html_file, 'html.parser')
            print('data collected')
        self.window['-OUTPUT-'].update(self.page.prettify())

    def request_data(self, source, pages):
        try:
            session = requests.Session()
            session.headers['User-Agent'] = self.USER_AGENT_HEADER
            response = ''
            for page in range(1, pages+1):
                if page != 1:
                    url = f'{source}{self.PAGE_URL_POSTFIX}{page}/'
                    sleep(self.PARSING_DELAY_TIME)
                else:
                    url = source
                response += session.get(url).text
            self.page = bs(response, 'html.parser')
            
            print('data collected')
            self.window['-OUTPUT-'].update(self.page.prettify())

        except Exception as e:
            print(f'an error occured:\n{e}')
        
    def find_data(self, container, attrs):
        if not self.page:
            print('page is not loaded')
            return

        self.data = self.page.find_all(container, attrs=attrs)
        result = ''
        for obj in self.data:
            result += obj.prettify()
        print(f'{len(self.data)} objects are found')
        self.window['-OUTPUT-'].update(result)
        
        if not self.data:
            print('objects are not found')
            self.window['-OUTPUT-'].update(self.page.prettify())

    def parse_data(self, params):
        if not self.data:
            print('no data to parse')
            return

        if not params:
            print('no parameters are set')
            return

        self.clean_data.clear()
        for obj in self.data:
            buff = {}
            for param in params:
                try:
                    value = obj.find(param['container'], attrs=param['attrs']).text.strip()
                except AttributeError:
                    value = None
                buff[param['name']] = value
            self.clean_data.append(buff)
        
        print('data parsed')
        self.window['-OUTPUT-'].update(self.clean_data)

    def load_conf(self, path):
        with open(path, mode='r', encoding='utf-8') as file:
            conf = json.load(file)
            [self.window[key].update(value) for key, value in conf.items()]
            print('conf loaded')

    def save_json(self, data, path):
        path = path + '.json' if not path.endswith('.json') else path
        with open(path, mode='w', encoding='utf-8') as file:
            json.dump(data, file, ensure_ascii=False, indent=4)
            print('data saved')

    def save_xlsx(self):
        if not self.clean_data:
            print('no parsed data to save')
            return
        
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        # Save keys in the first row
        for column, key in enumerate(self.clean_data[0].keys()):
            sheet.cell(row=1, column=column+1).value = key

        # Save values in the next rows 
        for row, obj in enumerate(self.clean_data):
            for column, value in enumerate(obj.values()):
                sheet.cell(row=row+2, column=column+1).value = value

        path = GUI.popup_get_file('Save as', no_window=True, save_as=True)
        if path:
            path = f'{path}.xlsx' if not path.endswith('.xlsx') else path
            workbook.save(path)
            print('data saved')
        
    def clear_param(self, param_number):
        self.window[f'-PARAM_CONTAINER_{param_number}-'].update('')
        self.window[f'-PARAM_SELECTOR_TYPE_{param_number}-'].update('')
        self.window[f'-PARAM_SELECTOR_NAME_{param_number}-'].update('')
        self.window[f'-PARAM_NAME_{param_number}-'].update('')

    def clear_data(self):
        self.data = []
        self.clean_data = []
        self.window['-MESSAGE-'].update('')
        self.window['-OUTPUT-'].update('')
        self.window['-PARSE-'].update(disabled=True)
        self.window['-SAVE_JSON-'].update(disabled=True)
        self.window['-SAVE_XLSX-'].update(disabled=True)
    

if __name__ == '__main__':
    parser = Parser()
    parser.run()
    parser.window.close()
    sys.exit()
